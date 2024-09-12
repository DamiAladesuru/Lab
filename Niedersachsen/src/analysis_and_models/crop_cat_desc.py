# %%
import pickle
import geopandas as gpd
import pandas as pd
import os
import math as m
from functools import reduce # For merging multiple DataFrames
import logging
import numpy as np
from shapely.geometry import box

# Set the current working directory
os.chdir('C:/Users/aladesuru/Documents/DataAnalysis/Lab/Niedersachsen')
# %%
from src.data import dataload as dl
from src.data import eca_new as eca

# additional gld columns
def bbox_width(geometry):
    minx, miny, maxx, maxy = geometry.bounds
    return maxx - minx

def bbox_length(geometry):
    minx, miny, maxx, maxy = geometry.bounds
    return maxy - miny

def square_cpar(gld): #shape index adjusted for square fields
    gld['cpar2'] = ((0.25 * gld['peri_m']) / (gld['area_m2']**0.5))
    return gld

def field_polspy(gld):
    gld['polspy'] = (gld['area_m2']) / ((gld['peri_m'] / 4)** 2)
    return gld

def bounding_box(gld):
    gld['bbox'] = gld['geometry'].apply(lambda geom: box(*geom.bounds))   
    return gld

def bbox_dimensions(gld):
    gld['bbox_width'] = gld['geometry'].apply(bbox_width)
    gld['bbox_length'] = gld['geometry'].apply(bbox_length)
    return gld

# %%
def add_missing_year_data(df, cellcode, from_year, to_year):
    # Filter the rows for the specified CELLCODE and from_year
    filtered_rows = df[(df['CELLCODE'] == cellcode) & (df['year'] == from_year)]
    
    # Create a copy of the filtered rows and update the year to to_year
    new_rows = filtered_rows.copy()
    new_rows['year'] = to_year
    
    # Concatenate the new rows to the original DataFrame
    df = pd.concat([df, new_rows], ignore_index=True)
    
    return df

def load_gld():
        # Load base data
    gld = dl.load_data(loadExistingData=True)
    # add additional columns to the data
    gld = square_cpar(gld)
    kulturcode_mastermap = eca.process_kulturcode()
    gld = pd.merge(gld, kulturcode_mastermap, on='kulturcode', how='left')
    gld = gld.drop(columns=['sourceyear'])
    # call the function to add missing year data
    gld = add_missing_year_data(gld, '10kmE438N336', 2016, 2017) 
    
    return gld


# %%
def create_catdf(gld):
    def compute_gridpar(df):
        # Sum of field size per grid (m2)
        fsm2_sum = gld.groupby(['year', 'CELLCODE', 'category2'])['area_m2'].sum().reset_index()
        fsm2_sum.columns = ['year', 'CELLCODE', 'category2', 'fsm2_sum']

        # Sum of field perimeter per grid
        peri_sum = gld.groupby(['year', 'CELLCODE', 'category2'])['peri_m'].sum().reset_index()
        peri_sum.columns = ['year', 'CELLCODE', 'category2', 'peri_sum']

        # Merge the two DataFrames
        df = pd.merge(fsm2_sum, peri_sum, on=['year', 'CELLCODE', 'category2'])

        # p/a ratio of grid as sum of peri divided by sum of area per grid
        df['grid_par'] = df['peri_sum'] / df['fsm2_sum']

        # Calculate the means based on year and category2
        means_df = df.groupby(['year', 'category2'])['grid_par'].mean().reset_index()
        
        return means_df

    ######################################################################
    columns = ['year', 'category2']

    # Extract the specified columns and remove duplicates
    catdf = gld[columns].drop_duplicates().copy()
    
    # Number of fields per grid
    fields = gld.groupby(['year', 'category2'])['geometry'].count().reset_index()
    fields.columns = ['year', 'category2', 'fields']
    catdf = pd.merge(catdf, fields, on=['year', 'category2'])
    
    # Sum of field size per grid
    fsha_sum = gld.groupby(['year', 'category2'])['area_ha'].sum().reset_index()
    fsha_sum.columns = ['year', 'category2', 'fsha_sum']
    catdf = pd.merge(catdf, fsha_sum, on=['year', 'category2'])

    # Mean field size per grid
    catdf['mfs_ha'] = (catdf['fsha_sum'] / catdf['fields'])

    # Rate of fields per hectare of land per grid
    catdf['fields_ha'] = (catdf['fields'] / catdf['fsha_sum'])
    
    ######################################################################
    #Shape
    ######################################################################
    # p/a ratio of grid as sum of peri divided by sum of area per grid
    gridpar = compute_gridpar(gld)
    catdf = pd.merge(catdf, gridpar, on=['year', 'category2'])
    
    # Sum of par per grid
    par_sum = gld.groupby(['year', 'category2'])['par'].sum().reset_index()
    par_sum.columns = ['year', 'category2', 'par_sum']
    catdf = pd.merge(catdf, par_sum, on=['year', 'category2'])

    # Mean par per grid
    catdf['mean_par'] = (catdf['par_sum'] / catdf['fields'])
    
    return catdf    
      
catdf = create_catdf(gld)


# %%
def calculate_differences(df): #yearly gridcell differences and differences from first year
    # Create a copy of the original DataFrame to avoid altering the original data
    griddf_ext = df.copy()
    
    # Ensure the data is sorted by 'LANDKREIS', 'CELLCODE', 'category2', and 'year'
    griddf_ext.sort_values(by=['category2', 'year'], inplace=True)
    numeric_columns = griddf_ext.select_dtypes(include='number').columns

    # Create a dictionary to store the new columns
    new_columns = {}

    # Calculate yearly difference for numeric columns and store in the dictionary
    for col in numeric_columns:
        new_columns[f'{col}_yearly_diff'] = griddf_ext.groupby(['category2'])[col].diff().fillna(0)

    # Filter numeric columns to exclude columns with '_yearly_diff'
    numeric_columns_no_diff = [col for col in numeric_columns if not col.endswith('_yearly_diff')]

    # Calculate difference relative to the first year
    y1_df = griddf_ext.groupby(['category2']).first().reset_index()
    
    # Rename the numeric columns to indicate the first year
    y1_df = y1_df[['category2'] + list(numeric_columns_no_diff)]
    y1_df = y1_df.rename(columns={col: f'{col}_y1' for col in numeric_columns_no_diff})

    # Merge the first year values back into the original DataFrame
    griddf_ext = pd.merge(griddf_ext, y1_df, on=['category2'], how='left')

    # Calculate the difference from the first year for each numeric column (excluding yearly differences)
    for col in numeric_columns_no_diff:
        new_columns[f'{col}_diff_from_y1'] = griddf_ext[col] - griddf_ext[f'{col}_y1']
        new_columns[f'{col}_percdiff_to_y1'] = ((griddf_ext[col] - griddf_ext[f'{col}_y1']) / griddf_ext[f'{col}_y1'])*100

    # Drop the temporary first year columns
    griddf_ext.drop(columns=[f'{col}_y1' for col in numeric_columns_no_diff], inplace=True)

    # Concatenate the new columns to the original DataFrame all at once
    new_columns_df = pd.DataFrame(new_columns)
    griddf_ext = pd.concat([griddf_ext, new_columns_df], axis=1)

    return griddf_ext 

cat_averages = calculate_differences(catdf)

###########################################################################
# %%
# Remove rows belonging to 'sonstige flächen'
cat_averages = cat_averages[cat_averages['category2'] != 'sonstige flächen']

# %%
# Define the output directory
output_dir = 'reports/figures/plots/new'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Category2 dictionary color mapping
hue_colors = {
    'getreide': '#bcbd22',
    'ackerfutter': '#8c564b',
    'dauergrünland': '#006D5B',
    'gemüse': '#d62728',
    'hackfrüchte': '#9467bd',
    'ölsaaten': '#e377c2',
    'leguminosen': '#ff7f0e',
    'mischkultur': '#17becf',
    'dauerkulturen': '#2ca02c',

    'environmental': '#1f77b4'
}    #'sonstige flächen': '#7f7f7f',

# Translation dictionary
translation_dict = {
    'ackerfutter': 'forage crops',
    'gemüse': 'vegetables',
    'environmental': 'environmental areas',
    'leguminosen': 'legumes',
    'mischkultur': 'mixed culture',
    'hackfrüchte': 'root crops',
    'dauergrünland': 'permanent grassland',
    'ölsaaten': 'oilseeds',
    'dauerkulturen': 'perennial crops',
    'getreide': 'cereals',
    'sonstige flächen': 'other areas'
}

# %% Show the colors for each category in category2
plt.figure(figsize=(12, 6))
for category, color in hue_colors.items():
    plt.plot([], [], color=color, label=translation_dict[category], marker='o', linestyle='')

# Set the legend to be in two rows with a custom background color and increased border padding
ncol = math.ceil(len(hue_colors) / 2)
legend = plt.legend(title='Category2', bbox_to_anchor=(0.5, -0.1), loc='upper center', ncol=ncol, borderpad=2)
plt.axis('off')

# Adjust layout to prevent cut-off
plt.tight_layout()

# Save plot with bbox_inches='tight' to include all elements
plt.savefig(os.path.join(output_dir, 'category2_legend.png'), bbox_inches='tight')

plt.show()

# %% 
# fig 1: multi-line plot showing total land area for each crop group over time
#################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Create a line plot for each category with custom colors
sns.lineplot(data=cat_averages, x='year', y='fsha_sum', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Total Land Area (ha) for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('Total Land Area (ha)')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot
#plt.savefig(os.path.join(output_dir, 'tla_category2.svg'))

# Show the plot
plt.show()

# %%
# fig 2: multi-line plot showing average diff from base year of total agricultural land for each crop group over time
#######################################################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Set the background color
#plt.gca().set_facecolor('#e6e6e6')
#plt.gcf().set_facecolor('#e6e6e6')

# Create a line plot for each category with custom colors
sns.lineplot(data = cat_averages, x='year', y='fsha_sum_percdiff_to_y1', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Relative Difference from Year One of TL (ha) for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('R. Diff of TL (ha) from Year One')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot before showing it
plt.savefig(os.path.join(output_dir, 'diff_y1_TL_cat2.svg'))

# Show the plot
plt.show()


# %% 
# fig 3: multi-line plot showing average diff from base year of mean field size for each crop group over time
############################################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Set the background color
#plt.gca().set_facecolor('#e6e6e6')
#plt.gcf().set_facecolor('#e6e6e6')

# Create a line plot for each category with custom colors
sns.lineplot(data=cat_averages, x='year', y='mfs_ha_percdiff_to_y1', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Relative Difference from Year One of MFS (ha) for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('R. Diff of MFS (ha) from Year One')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot before showing it
plt.savefig(os.path.join(output_dir, 'diff_y1_MFS_cat2_oo.svg'))

# Show the plot
plt.show()

# %% fig 4: multi-line plot showing av. diff from base year of mean par for each crop group over time
########################################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Create a line plot for each category with custom colors
sns.lineplot(data=cat_averages, x='year', y='mean_par_percdiff_to_y1', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Difference from Year One of Average Mean PAR for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('R. Diff of Average MeanPAR from Year One')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot before showing it
plt.savefig(os.path.join(output_dir, 'diff_y1_amPAR_cat2_oo.svg'))

# Show the plot
plt.show()

# %% fig 5: multi-line plot showing av diff from base year of GRID par for each crop group over time
########################################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Set the background color
#plt.gca().set_facecolor('#e6e6e6')
#plt.gcf().set_facecolor('#e6e6e6')

# Create a line plot for each category with custom colors
sns.lineplot(data=cat_averages, x='year', y='grid_par_percdiff_to_y1', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Relative Difference from Year One of Avergae Grid PAR for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('R. Difference of Average Grid PAR from Year One')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot before showing it
plt.savefig(os.path.join(output_dir, 'diff_y1_aGPAR_cat2_oo.svg'))

# Show the plot
plt.show()

# %% fig 6: multi-line plot showing av diff from base year of fields_ha for each crop group over time
########################################################################################################
# Set the plot style
sns.set(style="whitegrid")

# Create a figure
plt.figure(figsize=(12, 6))

# Set the background color
#plt.gca().set_facecolor('#e6e6e6')
#plt.gcf().set_facecolor('#e6e6e6')

# Create a line plot for each category with custom colors
sns.lineplot(data=cat_averages, x='year', y='fields_ha_percdiff_to_y1', hue='category2',
             marker='o', palette=hue_colors, legend=False)

# Add titles and labels
plt.title('Trend of Relative Difference from Year One of Fields/ha for Each Crop Group Over Time')
plt.xlabel('Year')
plt.ylabel('R. Diff of Fields_ha from Year One')
#plt.legend(title='Crop Group', bbox_to_anchor=(1.05, 1), loc='right')

# Remove the top and right spines
sns.despine(left=True, bottom=True)

# Save plot before showing it
plt.savefig(os.path.join(output_dir, 'diff_y1_fieldsha_cat2_oo.svg'))

# Show the plot
plt.show()

# %%
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# Define the hue colors
hue_colors = {
    'getreide': '#bcbd22',
    'ackerfutter': '#8c564b',
    'dauergrünland': '#006D5B',
    'gemüse': '#d62728',
    'hackfrüchte': '#9467bd',
    'ölsaaten': '#e377c2',
    'leguminosen': '#ff7f0e',
    'mischkultur': '#17becf',
    'dauerkulturen': '#2ca02c',
    'sonstige flächen': '#7f7f7f',
    'environmental': '#1f77b4',
}

# Write the DataFrame to an Excel file
excel_path = 'cat_averages.xlsx'
cat_averages.to_excel(excel_path, index=False)

# Load the workbook and select the active sheet
wb = load_workbook(excel_path)
ws = wb.active

# Apply the color fill to the 'category2' column based on its values
category2_col_idx = cat_averages.columns.get_loc('category2') + 1  # Get the column index for 'category2'

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=category2_col_idx, max_col=category2_col_idx):
    for cell in row:
        category = cell.value
        if category in hue_colors:
            fill_color = hue_colors[category]
            cell.fill = PatternFill(start_color=fill_color[1:], end_color=fill_color[1:], fill_type="solid")

# Save the workbook
wb.save(excel_path)

# %%
import matplotlib.pyplot as plt

# Group the DataFrame by year and sum the 'fsha_sum' values
fsha_sum_by_year = cat_averages.groupby('year')['fsha_sum'].sum().reset_index()

# Plot the summed values
plt.figure(figsize=(10, 6))
plt.plot(fsha_sum_by_year['year'], fsha_sum_by_year['fsha_sum'], marker='o', linestyle='-', color='b')
plt.xlabel('Year')
plt.ylabel('Total fsha_sum')
plt.title('Total fsha_sum by Year')
plt.grid(True)
plt.show()
# %%
